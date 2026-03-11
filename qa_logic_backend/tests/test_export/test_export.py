import io
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.schemas.response import GenerationResponse, GeneratedTestCase, ResponseMetadata

client = TestClient(app)

def create_mock_tc(id: str, severity: int, probability: int, linked_req: str = "REQ-1"):
    return GeneratedTestCase(
        id=id,
        linked_requirement=linked_req,
        title=f"Test {id}",
        preconditions="Pre 1\nPre 2",
        steps=["Step 1", "Step 2"],
        expected_result="Pass",
        priority="High",
        category="Positive",
        severity=severity,
        probability=probability,
        risk_score=severity * probability,
        risk_level="High",
        notes="note 1\nnote 2"
    )

def create_mock_export_payload() -> dict:
    cases = [create_mock_tc("TC-1", 5, 5)]
    response = GenerationResponse(
        metadata=ResponseMetadata(truncated=False, original_count=1),
        test_cases=cases,
        clarification_questions=[],
        assumptions=[]
    )
    # Cast via `.model_dump()` to represent JSON natively sent from Client
    return {"data": response.model_dump()}

def test_export_invalid_payload_fails():
    '''Missing exclusively required 'data' or 'history_id' returns 422.'''
    response = client.post("/api/v1/export", json={})
    assert response.status_code == 422

def test_export_history_id_not_implemented():
    '''Verifies history retrieval returns explicit expected 501 bounds natively.'''
    response = client.post("/api/v1/export", json={"history_id": "historical-uuid"})
    assert response.status_code == 501
    assert "Phase 6" in response.json()["detail"]

def test_export_success_workbook():
    '''Ensures payload properly structures to Excel in-memory and streams cleanly.'''
    payload = create_mock_export_payload()
    
    response = client.post("/api/v1/export", json=payload)
    
    # 1. Check structural headers
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=\"qa-logic-pro-test-cases-" in response.headers["content-disposition"]
    
    # 2. Check workbook structural composition natively (using bytes buffer intercept)
    buffer = io.BytesIO(response.content)
    wb = load_workbook(buffer)
    
    assert "Test Cases" in wb.sheetnames
    assert "Risk & Notes" in wb.sheetnames
    
    ws_tests = wb["Test Cases"]
    
    # 3. Check specific cells natively
    assert ws_tests["A1"].value == "ID"
    assert ws_tests["K1"].value == "Steps"
    
    # Check steps array merged natively preserving real Windows/Excel newline rendering
    steps_val = ws_tests["K2"].value
    assert "Step 1\nStep 2" in steps_val
    
    # Ensure notes are loaded properly in Risk mapping
    ws_risk = wb["Risk & Notes"]
    
    # 4. Check for existence of the new structural headers in Risk & Notes
    found_summary = False
    found_notes = False
    
    notes_val = None
    
    for row in ws_risk.iter_rows(values_only=True):
        if row[0] == "Risk Summary":
            found_summary = True
        if row[0] == "TC-1":
            notes_val = row[3]
            
    assert found_summary, "Risk Summary section missing from Risk & Notes sheet"
    assert notes_val is not None, "Test Case notes not found in Risk & Notes sheet"
    assert "note 1\nnote 2" in notes_val
