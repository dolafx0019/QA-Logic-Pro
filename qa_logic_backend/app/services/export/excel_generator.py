import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from app.schemas.response import GenerationResponse

def generate_excel_bytes(response_data: GenerationResponse) -> io.BytesIO:
    """
    Generates an in-memory .xlsx representation of the GenerationResponse payload.
    It strictly builds the workbook from the provided backend schema deterministically.
    """
    wb = Workbook()
    
    # 1. Sheet 1: Test Cases
    ws_tests = wb.active
    ws_tests.title = "Test Cases"
    
    headers_1 = [
        "ID", "Linked Requirement", "Title", "Category", "Priority", 
        "Severity", "Probability", "Risk Score", "Risk Level", 
        "Preconditions", "Steps", "Expected Result"
    ]
    ws_tests.append(headers_1)
    
    # Format Headers
    header_font = Font(bold=True)
    for cell in ws_tests[1]:
        cell.font = header_font
        
    for tc in response_data.test_cases:
        # Join multiline fields with real newlines for Excel
        steps_text = "\n".join(tc.steps) if tc.steps else ""
        row = [
            tc.id,
            tc.linked_requirement,
            tc.title,
            tc.category,
            tc.priority,
            tc.severity,
            tc.probability,
            tc.risk_score,
            tc.risk_level,
            tc.preconditions,
            steps_text,
            tc.expected_result
        ]
        ws_tests.append(row)
        
    # Styles / Wrap text
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    max_column_width = 50
    
    # Apply wrap to long-text columns (Preconditions=10, Steps=11, Expected=12)
    # openpyxl uses 1-based indexing if accessing iter_cols but row[...] is 0-based for lists.
    # ws.iter_rows gives Cell objects. Here row is a tuple of Cells.
    for row_cells in ws_tests.iter_rows(min_row=2):
        row_cells[9].alignment = wrap_alignment  # Preconditions
        row_cells[10].alignment = wrap_alignment # Steps
        row_cells[11].alignment = wrap_alignment # Expected Result
            
    # Auto-adjust column widths (bounded)
    for col in ws_tests.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # Find longest line in multiline strings
                lines = str(cell.value).split('\n')
                longest = max((len(l) for l in lines), default=0)
                if longest > max_len:
                    max_len = longest
        # Apply reasonable bounding wrapper max width
        adjusted = min(max_len + 2, max_column_width)
        ws_tests.column_dimensions[col_letter].width = adjusted
        
    # 2. Sheet 2: Risk & Notes
    ws_risk = wb.create_sheet(title="Risk & Notes")
    
    # 2a. Summary Section
    ws_risk.append(["Risk Summary"])
    ws_risk["A1"].font = header_font
    
    if response_data.risk_summary:
        ws_risk.append(["Total High:", response_data.risk_summary.total_high])
        ws_risk.append(["Total Medium:", response_data.risk_summary.total_medium])
        ws_risk.append(["Total Low:", response_data.risk_summary.total_low])
        ws_risk.append(["Average Score:", response_data.risk_summary.average_score])
    else:
        ws_risk.append(["No risk summary available."])
        
    ws_risk.append([]) # Blank row
    
    # 2b. Assumptions Section
    ws_risk.append(["Assumptions"])
    ws_risk.cell(row=ws_risk.max_row, column=1).font = header_font
    if response_data.assumptions:
        ws_risk.append(["Assumption", "Rationale"])
        for cell in ws_risk[ws_risk.max_row]:
            cell.font = header_font
        for item in response_data.assumptions:
            ws_risk.append([item.assumption, item.rationale])
    else:
        ws_risk.append(["No assumptions provided."])
        
    ws_risk.append([]) # Blank row
    
    # 2c. Clarification Questions Section
    ws_risk.append(["Clarification Questions"])
    ws_risk.cell(row=ws_risk.max_row, column=1).font = header_font
    if response_data.clarification_questions:
        ws_risk.append(["Question", "Context"])
        for cell in ws_risk[ws_risk.max_row]:
            cell.font = header_font
        for cq in response_data.clarification_questions:
            ws_risk.append([cq.question, cq.context])
    else:
        ws_risk.append(["No clarification questions provided."])
        
    ws_risk.append([]) # Blank row

    # 2d. Test Cases Risk Notes Section
    ws_risk.append(["Test Cases Risk & Notes List"])
    ws_risk.cell(row=ws_risk.max_row, column=1).font = header_font
    
    ws_risk.append(["Test ID", "Requirement", "Risk Level", "Notes"])
    for cell in ws_risk[ws_risk.max_row]:
        cell.font = header_font
        
    for tc in response_data.test_cases:
        ws_risk.append([
            tc.id, tc.linked_requirement, tc.risk_level, tc.notes
        ])
        
    for row_cells in ws_risk.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = wrap_alignment
        
    ws_risk.column_dimensions['A'].width = 30
    ws_risk.column_dimensions['B'].width = 40
    ws_risk.column_dimensions['C'].width = 20
    ws_risk.column_dimensions['D'].width = 60
    
    # 3. Save to In-Memory byte buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
